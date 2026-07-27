"""Generate the effort/time estimation workbook — a SINGLE sheet with the plan
as effort line items (both milestones), totals, and achievable accuracy.

Effort is in working days; time in calendar weeks. No roles/team are named.

  PYTHONPATH=backend python -m scripts.make_estimation_xlsx
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = "Foot-Scan-Effort-Estimation.xlsx"
NAVY, TEAL, LIGHT, GREY, WHITE = "1F3A5F", "2E6E6A", "EAF1F4", "F4F6F7", "FFFFFF"
HOURS_PER_DAY = 8   # effort is authored in days, presented in hours

hdr = Font(bold=True, color=WHITE, size=11)
title = Font(bold=True, color=NAVY, size=16)
sub = Font(italic=True, color="555555", size=10)
bold = Font(bold=True, color=NAVY)
thin = Side(style="thin", color="C9D3D8")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center")

# (milestone, week, work_item, description, effort_days)   band rows have work_item=None
ROWS = [
    ("Milestone 1 — Full Production Pipeline (3 weeks)", None, None, None, None),
    ("M1", "W1", "Capture protocol + in-app guidance",
     "Enforce weight-bearing, ankle framing, scale marker/LiDAR, coverage, lighting; real-time capture hints.", 3),
    ("M1", "W1", "Ingestion + video keyframes",
     "Accept image batches and video; extract sharpest/most-diverse keyframes; normalise formats (HEIC).", 2),
    ("M1", "W1", "Image quality gate (Agent 1)",
     "Blur, exposure, foot presence, angle coverage/diversity, resolution, marker presence; reject reasons.", 3),
    ("M1", "W1", "Foot localisation (vision-LLM crop)",
     "Vision model locates the foot per frame; crop to it (keeps alignment texture, drops leg/clutter) so background can't fuse into the mesh.", 4),
    ("M1", "W2", "KIRI reconstruction integration",
     "Submit → webhook (signed) + poll fallback → fetch within link window; retries + credit metering.", 3),
    ("M1", "W2", "Mesh cleanup + watertight repair",
     "Vertex merge, largest-component isolation, hole repair (pymeshfix) before any ray casting.", 2),
    ("M1", "W2", "Scale calibration (marker / LiDAR)",
     "Reference-marker → mm/px, or LiDAR metric units. The one error nothing downstream can catch.", 3),
    ("M1", "W2", "Orientation (leg crop + stable-pose)",
     "Crop the leg at the ankle, physics stable-pose to rest the sole on a floor; adaptive fallback.", 3),
    ("M1", "W2", "Measurement + validation + reliability flags",
     "Length/width from footprint; arch by multi-slice ray casting; validate vs ground truth; flag implausible values.", 4),
    ("M1", "W3", "Standardized renders",
     "Deterministic camera poses + dimensioned overlays, headless.", 2),
    ("M1", "W3", "Vision-LLM biomechanics (Agent 2)",
     "Arch type, pronation/supination, estimated weight distribution → structured output.", 2),
    ("M1", "W3", "LangGraph orchestration",
     "Typed shared state, checkpointing, resumable runs, clinician review interrupt, audit trail.", 3),
    ("M1", "W3", "Automated report — baseline (Agent 3)",
     "Structured podiatry report (measurements + confidence + assessment + imagery).", 2),
    ("M1", "W3", "Clinician web app",
     "Intake, live progress, review/approve-edit, results + reliability flags + report view.", 4),
    ("M1", "W3", "End-to-end + accuracy validation",
     "Full run on real captures; validate against ground truth at the ±1 mm benchmark.", 2),

    ("Milestone 2 — Refinement & Improvement (2 weeks)", None, None, None, None),
    ("M2", "W4", "Parametric insole (.stl) generation (Agent 4)",
     "Parametric CAD from measurements + prescription → manufacturable 3D-printable STL + manifest.", 4),
    ("M2", "W4", "Foot-only segmentation upgrade",
     "Prompted/trained foot segmenter so the leg never enters reconstruction (fixes width on cluttered captures).", 3),
    ("M2", "W4", "Arch + measurement hardening",
     "Multi-scan/multi-slice averaging, arch refinement, medial/lateral robustness; tighten reliability thresholds.", 3),
    ("M2", "W5", "Clinical validation set",
     "Validate against additional ground-truth feet; publish per-measurement accuracy report.", 2),
    ("M2", "W5", "Errors, retries, edge cases",
     "Reconstruction failures, timeouts, missed webhooks, bad captures; graceful degradation.", 2),
    ("M2", "W5", "Report polish (PDF) + manifest",
     "Professional PDF report; manufacturing spec/manifest export.", 2),
    ("M2", "W5", "Observability + cost + deploy hardening",
     "Per-node metrics, error tracking, KIRI/LLM cost dashboards; deployment + data-security/compliance.", 3),
]

ACCURACY = [
    ("Foot length / width", "±1–2 mm, validated at the ±1 mm clinical benchmark"),
    ("Arch height", "Reliable from weight-bearing capture"),
    ("Scale", "Metric (reference marker or LiDAR)"),
    ("Insole output", "Manufacturable 3D-printable .stl (end of Milestone 2)"),
    ("Trust", "Reliability flags on every measurement; implausible values are flagged, never shipped"),
]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan & Effort"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Foot Scan → Custom Insole — Plan & Effort Estimation"
    ws["A1"].font = title
    ws["A2"] = ("One sheet. Effort in hours (8 h/day); time in calendar weeks. "
                "Milestone 1 = 3 weeks (full pipeline), Milestone 2 = 2 weeks (refinement).")
    ws["A2"].font = sub

    headers = ["Milestone", "Week", "Work Item", "Description", "Effort (hours)"]
    for c, t in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=t)
        cell.font = hdr; cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = center; cell.border = border

    r = 5
    m1 = m2 = 0
    for ms, wk, item, desc, eff in ROWS:
        if item is None:  # milestone band
            cell = ws.cell(row=r, column=1, value=ms)
            cell.font = Font(bold=True, color=WHITE, size=12)
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=TEAL)
                ws.cell(row=r, column=c).border = border
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            r += 1
            continue
        ws.cell(row=r, column=1, value=ms).font = bold
        ws.cell(row=r, column=2, value=wk)
        ws.cell(row=r, column=3, value=item).font = bold
        ws.cell(row=r, column=4, value=desc)
        ws.cell(row=r, column=5, value=eff * HOURS_PER_DAY)
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = border
            ws.cell(row=r, column=c).alignment = wrap
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=GREY)
        ws.cell(row=r, column=1).alignment = center
        ws.cell(row=r, column=2).alignment = center
        ws.cell(row=r, column=5).alignment = center
        m1 += eff * HOURS_PER_DAY if ms == "M1" else 0
        m2 += eff * HOURS_PER_DAY if ms == "M2" else 0
        r += 1

    # totals
    for label, val, fill in [("Milestone 1 total (3 weeks) · hours", m1, NAVY),
                             ("Milestone 2 total (2 weeks) · hours", m2, NAVY),
                             ("TOTAL (5 weeks) · hours", m1 + m2, "111C2B")]:
        ws.cell(row=r, column=3, value=label).font = bold
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="right")
        c5 = ws.cell(row=r, column=5, value=val)
        c5.font = Font(bold=True, color=WHITE); c5.alignment = center
        c5.fill = PatternFill("solid", fgColor=fill)
        r += 1

    # achievable accuracy
    r += 2
    ws.cell(row=r, column=1, value="Achievable accuracy (with the mandated capture protocol)").font = title
    r += 1
    for c, t in enumerate(["Aspect", "Outcome"], 1):
        cell = ws.cell(row=r, column=c, value=t)
        cell.font = hdr; cell.fill = PatternFill("solid", fgColor=TEAL); cell.border = border; cell.alignment = center
    r += 1
    for aspect, outcome in ACCURACY:
        ws.cell(row=r, column=1, value=aspect).font = bold
        ws.cell(row=r, column=2, value=outcome)
        for c in (1, 2):
            ws.cell(row=r, column=c).border = border; ws.cell(row=r, column=c).alignment = wrap
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Note").font = bold
    ws.cell(row=r, column=2, value="Accuracy depends on capture quality. The capture protocol "
            "(weight-bearing, foot-framed, scale marker/LiDAR, coverage) is a prerequisite, enforced by the quality gate.")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    ws.cell(row=r, column=2).alignment = wrap

    for i, w in enumerate([12, 8, 34, 60, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUT)
    print(f"wrote {OUT} (single sheet) — M1={m1}h, M2={m2}h, total={m1+m2}h")


if __name__ == "__main__":
    main()
